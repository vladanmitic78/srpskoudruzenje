"""
Attendance Report Generator
Generates PDF and Excel reports for attendance statistics
"""

import os
from datetime import datetime, timedelta
from io import BytesIO
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import mm, cm
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, PageBreak
from reportlab.lib.enums import TA_CENTER, TA_RIGHT, TA_LEFT
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont

# Try to register DejaVu fonts for Serbian character support
DEJAVU_FONT_PATH = "/usr/share/fonts/truetype/dejavu/"
try:
    pdfmetrics.registerFont(TTFont('DejaVuSans', f'{DEJAVU_FONT_PATH}DejaVuSans.ttf'))
    pdfmetrics.registerFont(TTFont('DejaVuSans-Bold', f'{DEJAVU_FONT_PATH}DejaVuSans-Bold.ttf'))
    FONT_NORMAL = 'DejaVuSans'
    FONT_BOLD = 'DejaVuSans-Bold'
except:
    FONT_NORMAL = 'Helvetica'
    FONT_BOLD = 'Helvetica-Bold'

# Brand colors
PRIMARY_COLOR = colors.HexColor('#C1272D')
SUCCESS_COLOR = colors.HexColor('#28a745')
WARNING_COLOR = colors.HexColor('#ffc107')
DANGER_COLOR = colors.HexColor('#dc3545')


def generate_attendance_pdf_report(
    report_data: dict,
    output_path: str = None,
    report_title: str = "Izveštaj o Prisustvu / Attendance Report"
) -> bytes:
    """
    Generate a PDF attendance report
    
    Args:
        report_data: Dictionary containing:
            - summary: Overall statistics
            - events: List of events with attendance
            - members: List of members with their attendance stats
            - date_range: { start, end }
            - training_group: Optional filter
            - generated_at: Timestamp
            - generated_by: Admin name
        output_path: Optional file path to save PDF
        
    Returns:
        PDF bytes if no output_path, otherwise saves to file
    """
    
    buffer = BytesIO()
    
    doc = SimpleDocTemplate(
        buffer,
        pagesize=landscape(A4),
        rightMargin=1.5*cm,
        leftMargin=1.5*cm,
        topMargin=1.5*cm,
        bottomMargin=1.5*cm
    )
    
    elements = []
    styles = getSampleStyleSheet()
    
    # Custom styles
    title_style = ParagraphStyle(
        'ReportTitle',
        parent=styles['Heading1'],
        fontName=FONT_BOLD,
        fontSize=18,
        textColor=PRIMARY_COLOR,
        alignment=TA_CENTER,
        spaceAfter=5*mm
    )
    
    subtitle_style = ParagraphStyle(
        'Subtitle',
        parent=styles['Normal'],
        fontName=FONT_NORMAL,
        fontSize=10,
        textColor=colors.gray,
        alignment=TA_CENTER,
        spaceAfter=10*mm
    )
    
    section_header = ParagraphStyle(
        'SectionHeader',
        parent=styles['Heading2'],
        fontName=FONT_BOLD,
        fontSize=12,
        textColor=PRIMARY_COLOR,
        spaceBefore=8*mm,
        spaceAfter=4*mm
    )
    
    normal_style = ParagraphStyle(
        'CustomNormal',
        parent=styles['Normal'],
        fontName=FONT_NORMAL,
        fontSize=9,
        leading=12
    )
    
    # ===== TITLE =====
    elements.append(Paragraph(report_title, title_style))
    
    date_range = report_data.get('date_range', {})
    training_group = report_data.get('training_group', 'Sve grupe / All groups')
    generated_at = report_data.get('generated_at', datetime.utcnow().isoformat())
    generated_by = report_data.get('generated_by', 'Admin')
    
    elements.append(Paragraph(
        f"Period: {date_range.get('start', 'N/A')} - {date_range.get('end', 'N/A')} | "
        f"Grupa: {training_group}<br/>"
        f"Generisano: {generated_at[:10]} od {generated_by}",
        subtitle_style
    ))
    
    # ===== SUMMARY SECTION =====
    elements.append(Paragraph("📊 Ukupna Statistika / Overall Statistics", section_header))
    
    summary = report_data.get('summary', {})
    
    summary_data = [
        ['Ukupno Događaja\nTotal Events', 'Ukupno Članova\nTotal Members', 
         'Prosečno Prisustvo\nAvg Attendance', 'Ukupno Prisutnih\nTotal Present',
         'Ukupno Odsutnih\nTotal Absent', 'Walk-in'],
        [
            str(summary.get('total_events', 0)),
            str(summary.get('total_members', 0)),
            f"{summary.get('average_attendance_rate', 0):.1f}%",
            str(summary.get('total_present', 0)),
            str(summary.get('total_absent', 0)),
            str(summary.get('total_walkins', 0))
        ]
    ]
    
    summary_table = Table(summary_data, colWidths=[4.5*cm, 4.5*cm, 4.5*cm, 4.5*cm, 4.5*cm, 3*cm])
    summary_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), PRIMARY_COLOR),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, 0), FONT_BOLD),
        ('FONTSIZE', (0, 0), (-1, 0), 8),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('FONTNAME', (0, 1), (-1, -1), FONT_BOLD),
        ('FONTSIZE', (0, 1), (-1, -1), 14),
        ('BACKGROUND', (0, 1), (-1, -1), colors.HexColor('#f8f9fa')),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#dee2e6')),
        ('TOPPADDING', (0, 0), (-1, -1), 8),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
    ]))
    elements.append(summary_table)
    elements.append(Spacer(1, 8*mm))
    
    # ===== EVENTS BREAKDOWN =====
    elements.append(Paragraph("📅 Pregled po Događajima / Events Breakdown", section_header))
    
    events = report_data.get('events', [])
    
    if events:
        events_header = ['Datum\nDate', 'Događaj\nEvent', 'Grupa\nGroup', 
                        'Prijavljeni\nConfirmed', 'Prisutni\nPresent', 
                        'Odsutni\nAbsent', 'Walk-in', 'Stopa\nRate']
        events_data = [events_header]
        
        for event in events[:20]:  # Limit to 20 events per page
            attendance_rate = 0
            if event.get('confirmed', 0) > 0:
                attendance_rate = (event.get('present', 0) / event.get('confirmed', 1)) * 100
            
            events_data.append([
                event.get('date', 'N/A'),
                event.get('title', 'N/A')[:30],
                event.get('training_group', '-')[:15],
                str(event.get('confirmed', 0)),
                str(event.get('present', 0)),
                str(event.get('absent', 0)),
                str(event.get('walkin', 0)),
                f"{attendance_rate:.0f}%"
            ])
        
        events_table = Table(events_data, colWidths=[2.5*cm, 6*cm, 3*cm, 2.5*cm, 2.5*cm, 2.5*cm, 2*cm, 2*cm])
        events_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#343a40')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('FONTNAME', (0, 0), (-1, 0), FONT_BOLD),
            ('FONTSIZE', (0, 0), (-1, 0), 8),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('FONTNAME', (0, 1), (-1, -1), FONT_NORMAL),
            ('FONTSIZE', (0, 1), (-1, -1), 8),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f8f9fa')]),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#dee2e6')),
            ('TOPPADDING', (0, 0), (-1, -1), 5),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 5),
        ]))
        elements.append(events_table)
    else:
        elements.append(Paragraph("Nema događaja u izabranom periodu / No events in selected period", normal_style))
    
    elements.append(PageBreak())
    
    # ===== MEMBERS BREAKDOWN =====
    elements.append(Paragraph("👥 Pregled po Članovima / Members Breakdown", section_header))
    
    members = report_data.get('members', [])
    
    if members:
        members_header = ['Ime / Name', 'Email', 'Grupa\nGroup',
                         'Prijave\nRSVPs', 'Prisutni\nPresent', 
                         'Odsutni\nAbsent', 'Stopa\nRate']
        members_data = [members_header]
        
        # Sort by attendance rate
        sorted_members = sorted(members, key=lambda x: x.get('attendance_rate', 0), reverse=True)
        
        for member in sorted_members[:30]:  # Limit to 30 members per page
            members_data.append([
                member.get('name', 'N/A')[:25],
                member.get('email', 'N/A')[:30],
                member.get('training_group', '-')[:15],
                str(member.get('total_rsvps', 0)),
                str(member.get('total_present', 0)),
                str(member.get('total_absent', 0)),
                f"{member.get('attendance_rate', 0):.0f}%"
            ])
        
        members_table = Table(members_data, colWidths=[5*cm, 6*cm, 3*cm, 2.5*cm, 2.5*cm, 2.5*cm, 2*cm])
        members_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#343a40')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('FONTNAME', (0, 0), (-1, 0), FONT_BOLD),
            ('FONTSIZE', (0, 0), (-1, 0), 8),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('ALIGN', (0, 1), (1, -1), 'LEFT'),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('FONTNAME', (0, 1), (-1, -1), FONT_NORMAL),
            ('FONTSIZE', (0, 1), (-1, -1), 8),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f8f9fa')]),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#dee2e6')),
            ('TOPPADDING', (0, 0), (-1, -1), 5),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 5),
        ]))
        elements.append(members_table)
        
        # Top performers highlight
        if len(sorted_members) > 0:
            elements.append(Spacer(1, 5*mm))
            top_members = sorted_members[:5]
            elements.append(Paragraph("🏆 Top 5 po Prisustvu / Top 5 by Attendance", section_header))
            
            top_data = [['Rang', 'Ime / Name', 'Stopa Prisustva / Attendance Rate']]
            for i, m in enumerate(top_members, 1):
                top_data.append([f"#{i}", m.get('name', 'N/A'), f"{m.get('attendance_rate', 0):.1f}%"])
            
            top_table = Table(top_data, colWidths=[2*cm, 8*cm, 6*cm])
            top_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), SUCCESS_COLOR),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
                ('FONTNAME', (0, 0), (-1, -1), FONT_BOLD),
                ('FONTSIZE', (0, 0), (-1, -1), 10),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                ('BACKGROUND', (0, 1), (-1, -1), colors.HexColor('#d4edda')),
                ('GRID', (0, 0), (-1, -1), 0.5, SUCCESS_COLOR),
                ('TOPPADDING', (0, 0), (-1, -1), 6),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
            ]))
            elements.append(top_table)
    else:
        elements.append(Paragraph("Nema članova u izabranom periodu / No members in selected period", normal_style))
    
    # ===== FOOTER =====
    elements.append(Spacer(1, 10*mm))
    footer_style = ParagraphStyle(
        'Footer',
        parent=styles['Normal'],
        fontName=FONT_NORMAL,
        fontSize=8,
        textColor=colors.gray,
        alignment=TA_CENTER
    )
    elements.append(Paragraph("─" * 80, footer_style))
    elements.append(Paragraph(
        "Srpsko Kulturno Udruženje Täby | info@srpskoudruzenjetaby.se | www.srpskoudruzenjetaby.se",
        footer_style
    ))
    
    # Build PDF
    doc.build(elements)
    
    pdf_bytes = buffer.getvalue()
    buffer.close()
    
    if output_path:
        with open(output_path, 'wb') as f:
            f.write(pdf_bytes)
        return output_path
    
    return pdf_bytes


def generate_attendance_excel_report(report_data: dict) -> bytes:
    """
    Generate an Excel attendance report
    
    Returns:
        Excel file bytes
    """
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter
    
    wb = openpyxl.Workbook()
    
    # Styles
    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='C1272D', end_color='C1272D', fill_type='solid')
    header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # ===== SUMMARY SHEET =====
    ws_summary = wb.active
    ws_summary.title = "Pregled - Summary"
    
    summary = report_data.get('summary', {})
    date_range = report_data.get('date_range', {})
    
    ws_summary['A1'] = "Izveštaj o Prisustvu / Attendance Report"
    ws_summary['A1'].font = Font(bold=True, size=16, color='C1272D')
    ws_summary.merge_cells('A1:F1')
    
    ws_summary['A3'] = f"Period: {date_range.get('start', 'N/A')} - {date_range.get('end', 'N/A')}"
    ws_summary['A4'] = f"Grupa: {report_data.get('training_group', 'Sve grupe')}"
    ws_summary['A5'] = f"Generisano: {report_data.get('generated_at', '')[:10]}"
    
    # Summary stats
    summary_headers = ['Metrika / Metric', 'Vrednost / Value']
    summary_rows = [
        ['Ukupno Događaja / Total Events', summary.get('total_events', 0)],
        ['Ukupno Članova / Total Members', summary.get('total_members', 0)],
        ['Prosečno Prisustvo / Avg Attendance', f"{summary.get('average_attendance_rate', 0):.1f}%"],
        ['Ukupno Prisutnih / Total Present', summary.get('total_present', 0)],
        ['Ukupno Odsutnih / Total Absent', summary.get('total_absent', 0)],
        ['Walk-in', summary.get('total_walkins', 0)],
    ]
    
    start_row = 7
    for col, header in enumerate(summary_headers, 1):
        cell = ws_summary.cell(row=start_row, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
    
    for row_idx, row_data in enumerate(summary_rows, start_row + 1):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws_summary.cell(row=row_idx, column=col_idx, value=value)
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='center')
    
    ws_summary.column_dimensions['A'].width = 40
    ws_summary.column_dimensions['B'].width = 20
    
    # ===== EVENTS SHEET =====
    ws_events = wb.create_sheet("Događaji - Events")
    
    event_headers = ['Datum / Date', 'Događaj / Event', 'Grupa / Group', 
                     'Prijavljeni / Confirmed', 'Prisutni / Present', 
                     'Odsutni / Absent', 'Walk-in', 'Stopa / Rate']
    
    for col, header in enumerate(event_headers, 1):
        cell = ws_events.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
    
    events = report_data.get('events', [])
    for row_idx, event in enumerate(events, 2):
        attendance_rate = 0
        if event.get('confirmed', 0) > 0:
            attendance_rate = (event.get('present', 0) / event.get('confirmed', 1)) * 100
        
        row_data = [
            event.get('date', ''),
            event.get('title', ''),
            event.get('training_group', '-'),
            event.get('confirmed', 0),
            event.get('present', 0),
            event.get('absent', 0),
            event.get('walkin', 0),
            f"{attendance_rate:.0f}%"
        ]
        
        for col_idx, value in enumerate(row_data, 1):
            cell = ws_events.cell(row=row_idx, column=col_idx, value=value)
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='center')
    
    # Set column widths
    ws_events.column_dimensions['A'].width = 12
    ws_events.column_dimensions['B'].width = 35
    ws_events.column_dimensions['C'].width = 15
    for col in ['D', 'E', 'F', 'G', 'H']:
        ws_events.column_dimensions[col].width = 12
    
    # ===== MEMBERS SHEET =====
    ws_members = wb.create_sheet("Članovi - Members")
    
    member_headers = ['Ime / Name', 'Email', 'Grupa / Group',
                      'Prijave / RSVPs', 'Prisutni / Present', 
                      'Odsutni / Absent', 'Stopa / Rate']
    
    for col, header in enumerate(member_headers, 1):
        cell = ws_members.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
    
    members = report_data.get('members', [])
    sorted_members = sorted(members, key=lambda x: x.get('attendance_rate', 0), reverse=True)
    
    for row_idx, member in enumerate(sorted_members, 2):
        row_data = [
            member.get('name', ''),
            member.get('email', ''),
            member.get('training_group', '-'),
            member.get('total_rsvps', 0),
            member.get('total_present', 0),
            member.get('total_absent', 0),
            f"{member.get('attendance_rate', 0):.0f}%"
        ]
        
        for col_idx, value in enumerate(row_data, 1):
            cell = ws_members.cell(row=row_idx, column=col_idx, value=value)
            cell.border = thin_border
            if col_idx <= 2:
                cell.alignment = Alignment(horizontal='left')
            else:
                cell.alignment = Alignment(horizontal='center')
    
    ws_members.column_dimensions['A'].width = 25
    ws_members.column_dimensions['B'].width = 30
    ws_members.column_dimensions['C'].width = 15
    for col in ['D', 'E', 'F', 'G']:
        ws_members.column_dimensions[col].width = 12
    
    # Save to bytes
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    return buffer.getvalue()
